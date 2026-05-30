"""
Unified file text extraction.

Returns *structured* ParsedBlock objects (not just a flat string) so the
chunker can keep tables whole, attach page numbers, and preserve section
headings.

Routing (all free libraries):
  - .pdf            → pdfplumber  (text + tables + page numbers)
                       fallback → pdf2image + OpenCV + Tesseract for scans
  - .docx           → python-docx (paragraphs + tables + headings)
  - .xlsx / .xls    → openpyxl    (each sheet becomes a table block)
  - .png/.jpg/.jpeg → OpenCV preprocess + Tesseract OCR
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data types
# ---------------------------------------------------------------------------
@dataclass
class ParsedBlock:
    """A semantically meaningful piece of extracted content."""
    content: str
    block_type: str = "text"          # "text" | "table" | "heading"
    page_number: Optional[int] = None
    section_heading: Optional[str] = None


class UnsupportedFileTypeError(Exception):
    pass


class TextExtractionError(Exception):
    pass


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------
def extract_blocks(file_path: str | Path) -> Tuple[List[ParsedBlock], str]:
    """
    Parse a file into structured blocks.

    Returns:
        (blocks, file_type_label)
    Raises:
        UnsupportedFileTypeError, TextExtractionError
    """
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext == ".pdf":
        blocks = _extract_pdf(path)
        file_type = "pdf"
    elif ext == ".docx":
        blocks = _extract_docx(path)
        file_type = "docx"
    elif ext in {".xlsx", ".xls"}:
        blocks = _extract_excel(path)
        file_type = "excel"
    elif ext in {".png", ".jpg", ".jpeg"}:
        blocks = _extract_image(path)
        file_type = "image"
    else:
        raise UnsupportedFileTypeError(
            "This file type is not supported. Please upload PDF, Word, Excel, or image files."
        )

    # Strip empty / whitespace-only blocks
    blocks = [b for b in blocks if (b.content or "").strip()]
    if not blocks:
        raise TextExtractionError(
            "Could not extract any text from this file. Please verify the file is not "
            "corrupted, password-protected, or a low-quality scan."
        )
    return blocks, file_type


def extract_text(file_path: str | Path) -> Tuple[str, str]:
    """
    Legacy entry point for callers that still expect a flat string.
    Internally uses extract_blocks().
    """
    blocks, ft = extract_blocks(file_path)
    text = "\n\n".join(b.content for b in blocks)
    return text, ft


# ---------------------------------------------------------------------------
# Helpers shared across parsers
# ---------------------------------------------------------------------------
def _table_to_markdown(rows: list[list]) -> str:
    """Render a 2D array of cells as a Markdown table the LLM can read."""
    if not rows:
        return ""
    # Drop fully-empty rows
    cleaned = [
        [("" if c is None else str(c).strip()) for c in row]
        for row in rows
        if any((c is not None and str(c).strip()) for c in row)
    ]
    if not cleaned:
        return ""

    # Equalise column count
    max_cols = max(len(r) for r in cleaned)
    cleaned = [r + [""] * (max_cols - len(r)) for r in cleaned]

    header = cleaned[0]
    body = cleaned[1:] if len(cleaned) > 1 else []

    lines = ["| " + " | ".join(header) + " |"]
    lines.append("|" + "|".join(["---"] * max_cols) + "|")
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def _preprocess_image_for_ocr(pil_img):
    """OpenCV preprocessing for better OCR. Falls back to identity if cv2 missing."""
    try:
        import cv2
        import numpy as np
    except ImportError:
        return pil_img

    img = np.array(pil_img)
    if len(img.shape) == 3:
        gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
    else:
        gray = img

    # Light denoise
    try:
        denoised = cv2.fastNlMeansDenoising(gray, h=15)
    except Exception:
        denoised = gray

    # Otsu's threshold — sharpens text against background
    try:
        _, thresh = cv2.threshold(
            denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU
        )
        return thresh
    except Exception:
        return denoised


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _extract_pdf(path: Path) -> List[ParsedBlock]:
    """
    Extract text + tables from a digitally-generated PDF using pdfplumber.
    If the document has no extractable text (i.e., it's scanned), fall back
    to OCR via pdf2image + Tesseract.
    """
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber missing — falling back to PyMuPDF text-only mode.")
        return _extract_pdf_pymupdf(path)

    blocks: List[ParsedBlock] = []
    has_text = False

    try:
        with pdfplumber.open(str(path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                # --- Tables first ---
                try:
                    found_tables = page.find_tables() or []
                except Exception:
                    found_tables = []
                for tbl in found_tables:
                    try:
                        rows = tbl.extract()
                    except Exception:
                        rows = None
                    if not rows:
                        continue
                    md = _table_to_markdown(rows)
                    if md.strip():
                        blocks.append(
                            ParsedBlock(
                                content=md,
                                block_type="table",
                                page_number=page_num,
                            )
                        )

                # --- Text on the page (whole page, simple & robust) ---
                try:
                    text = page.extract_text() or ""
                except Exception:
                    text = ""
                text = text.strip()
                if text:
                    has_text = True
                    blocks.append(
                        ParsedBlock(
                            content=f"[Page {page_num}]\n{text}",
                            block_type="text",
                            page_number=page_num,
                        )
                    )
    except Exception as exc:
        logger.exception("pdfplumber failed: %s — trying PyMuPDF fallback.", exc)
        return _extract_pdf_pymupdf(path)

    # Likely a scanned PDF — fall back to OCR
    if not has_text and not any(b.block_type == "table" for b in blocks):
        logger.info("No extractable text in PDF — falling back to OCR.")
        ocr_blocks = _ocr_pdf(path)
        if ocr_blocks:
            return ocr_blocks

    return blocks


def _extract_pdf_pymupdf(path: Path) -> List[ParsedBlock]:
    """Legacy PyMuPDF fallback — text-only, no table preservation."""
    try:
        import fitz
    except ImportError as exc:
        raise RuntimeError("Neither pdfplumber nor PyMuPDF is installed") from exc

    blocks: List[ParsedBlock] = []
    with fitz.open(str(path)) as doc:
        for page_num, page in enumerate(doc, start=1):
            t = (page.get_text("text") or "").strip()
            if t:
                blocks.append(
                    ParsedBlock(
                        content=f"[Page {page_num}]\n{t}",
                        block_type="text",
                        page_number=page_num,
                    )
                )
    return blocks


def _ocr_pdf(path: Path) -> List[ParsedBlock]:
    """OCR a scanned PDF page by page."""
    try:
        from pdf2image import convert_from_path
        import pytesseract
    except ImportError as exc:
        logger.error("OCR fallback requires pdf2image + pytesseract (and poppler).")
        return []

    try:
        images = convert_from_path(str(path), dpi=200)
    except Exception as exc:
        logger.error("pdf2image failed (is poppler installed?): %s", exc)
        return []

    blocks: List[ParsedBlock] = []
    for page_num, img in enumerate(images, start=1):
        try:
            prep = _preprocess_image_for_ocr(img)
            text = (pytesseract.image_to_string(prep) or "").strip()
        except Exception as exc:
            logger.warning("OCR failed for page %d: %s", page_num, exc)
            text = ""
        if text:
            blocks.append(
                ParsedBlock(
                    content=f"[Page {page_num}]\n{text}",
                    block_type="text",
                    page_number=page_num,
                )
            )
    return blocks


# ---------------------------------------------------------------------------
# DOCX
# ---------------------------------------------------------------------------
def _extract_docx(path: Path) -> List[ParsedBlock]:
    try:
        from docx import Document
        from docx.text.paragraph import Paragraph
        from docx.table import Table
    except ImportError as exc:
        raise RuntimeError(
            "python-docx is not installed. Run: pip install python-docx"
        ) from exc

    doc = Document(str(path))
    blocks: List[ParsedBlock] = []
    current_heading: Optional[str] = None

    # Iterate body children in original document order
    body = doc.element.body
    for child in body.iterchildren():
        tag = child.tag.split("}")[-1]
        if tag == "p":
            para = Paragraph(child, doc)
            text = (para.text or "").strip()
            if not text:
                continue
            style = (para.style.name if para.style else "") or ""
            if style.lower().startswith("heading") or style.lower() == "title":
                # Track heading as context but don't emit it as its own searchable block
                current_heading = text
                blocks.append(
                    ParsedBlock(
                        content=text,
                        block_type="heading",
                        section_heading=text,
                    )
                )
            else:
                blocks.append(
                    ParsedBlock(
                        content=text,
                        block_type="text",
                        section_heading=current_heading,
                    )
                )
        elif tag == "tbl":
            table = Table(child, doc)
            rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
            md = _table_to_markdown(rows)
            if md.strip():
                blocks.append(
                    ParsedBlock(
                        content=md,
                        block_type="table",
                        section_heading=current_heading,
                    )
                )
    return blocks


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def _extract_excel(path: Path) -> List[ParsedBlock]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed") from exc

    blocks: List[ParsedBlock] = []
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows: list[list] = []
            for row in sheet.iter_rows(values_only=True):
                values = [
                    ("" if c is None else str(c).strip())
                    for c in row
                ]
                if any(v for v in values):
                    rows.append(values)
            if rows:
                md = _table_to_markdown(rows)
                if md.strip():
                    blocks.append(
                        ParsedBlock(
                            content=md,
                            block_type="table",
                            section_heading=f"Sheet: {sheet_name}",
                        )
                    )
    finally:
        wb.close()
    return blocks


# ---------------------------------------------------------------------------
# Image (OCR + table detection)
# ---------------------------------------------------------------------------
def _extract_image(path: Path) -> List[ParsedBlock]:
    try:
        import pytesseract
        from PIL import Image
    except ImportError as exc:
        raise RuntimeError("pytesseract / Pillow are not installed") from exc

    blocks: List[ParsedBlock] = []

    # Try table detection first
    try:
        from img2table.document import Image as I2TImage
        from img2table.ocr import TesseractOCR

        ocr = TesseractOCR(n_threads=1, lang="eng")
        i2t_doc = I2TImage(src=str(path))
        extracted_tables = i2t_doc.extract_tables(
            ocr=ocr,
            implicit_rows=True,
            borderless_tables=True,
            min_confidence=50,
        )
        for tbl in extracted_tables or []:
            try:
                df = tbl.df
                if df is None or df.empty:
                    continue
                rows = [list(df.columns)] + df.values.tolist()
                md = _table_to_markdown(rows)
                if md.strip():
                    blocks.append(
                        ParsedBlock(content=md, block_type="table", page_number=1)
                    )
            except Exception:
                continue
    except Exception as exc:
        logger.debug("img2table not available or failed (%s) — skipping table detection.", exc)

    # OCR the full image as text (always — tables above are bonus context)
    try:
        with Image.open(str(path)) as img:
            prep = _preprocess_image_for_ocr(img)
            text = (pytesseract.image_to_string(prep) or "").strip()
    except Exception as exc:
        logger.error("OCR failed for %s: %s", path, exc)
        text = ""

    if text:
        blocks.append(
            ParsedBlock(content=text, block_type="text", page_number=1)
        )

    return blocks
